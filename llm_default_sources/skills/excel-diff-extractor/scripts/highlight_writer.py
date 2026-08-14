"""Highlighted Excel diff writer.

Given a DiffResult and the path to the 'actual' file, produces a new .xlsx
where every mismatched cell is visually highlighted and the expected value
is attached as a cell comment. A Summary sheet is inserted at the front.

Color legend:
- Red fill       → cell value differs from expected
- Orange fill    → row/column/sheet only exists in one side
- Yellow fill    → structural issue (shape or column name mismatch)

Kept intentionally small: the aim is a glanceable diff, not a full reporting
suite. If the caller needs different styling, they can subclass or copy.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from comparator import DiffResult, SheetDiff

logger = logging.getLogger(__name__)

# Colors chosen for accessibility: high contrast on white, distinguishable
# for common forms of color-blindness.
FILL_DIFF = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_MISSING = PatternFill(start_color="FFD8A8", end_color="FFD8A8", fill_type="solid")
FILL_STRUCTURAL = PatternFill(
    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
)
FILL_HEADER = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

FONT_HEADER = Font(bold=True)
FONT_TITLE = Font(bold=True, size=14)


class HighlightedDiffWriter:
    """Writes a highlighted diff Excel file from a DiffResult."""

    def __init__(self, result: DiffResult) -> None:
        self.result = result

    def write(self, output_path: Path) -> Path:
        """Create the diff workbook at `output_path` and return that path."""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        # Remove the default empty sheet; we'll add our own.
        default_sheet = wb.active
        wb.remove(default_sheet)

        self._write_summary_sheet(wb)

        if self.result.file_type == "excel":
            self._write_excel_content_sheets(wb)
        elif self.result.file_type == "csv":
            self._write_csv_content_sheet(wb)
        # For "unsupported", the summary sheet tells the whole story.

        wb.save(output_path)
        logger.info("Wrote diff workbook: %s", output_path)
        return output_path

    # --- Summary sheet ------------------------------------------------------

    def _write_summary_sheet(self, wb: Workbook) -> None:
        # Slash is forbidden in Excel sheet names, so we can't use "Summary / 概要".
        ws = wb.create_sheet("Summary_概要", 0)

        ws["A1"] = "Comparison Summary / 比較サマリ"
        ws["A1"].font = FONT_TITLE
        ws.merge_cells("A1:D1")

        rows: list[tuple[str, Any]] = [
            ("Expected file / 期待値ファイル", self.result.expected_path),
            ("Actual file / 実際値ファイル", self.result.actual_path),
            ("File type / ファイル種別", self.result.file_type),
            ("Float tolerance / 浮動小数点許容誤差", self.result.tolerance),
            (
                "Status / ステータス",
                "✓ Identical / 一致"
                if self.result.is_identical
                else f"✗ {self.result.total_cell_differences} cell diff(s) / セル差分 {self.result.total_cell_differences} 件",
            ),
            ("Sheets compared / 比較シート数", len(self.result.sheet_diffs)),
        ]
        for i, (label, value) in enumerate(rows, start=3):
            ws.cell(row=i, column=1, value=label).font = FONT_HEADER
            ws.cell(row=i, column=2, value=value)

        if self.result.read_error:
            r = len(rows) + 4
            cell = ws.cell(row=r, column=1, value=f"Read error / 読み込みエラー: {self.result.read_error}")
            cell.fill = FILL_STRUCTURAL
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            return

        # Per-sheet breakdown
        header_row = len(rows) + 4
        headers = [
            "Sheet / シート",
            "Status / ステータス",
            "Cell diffs / セル差分",
            "Expected shape / 期待サイズ",
            "Actual shape / 実サイズ",
            "Notes / 備考",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER

        for offset, sd in enumerate(self.result.sheet_diffs, start=1):
            r = header_row + offset
            status, notes, fill = self._summarize_sheet(sd)
            ws.cell(row=r, column=1, value=sd.sheet_name)
            status_cell = ws.cell(row=r, column=2, value=status)
            if fill is not None:
                status_cell.fill = fill
            ws.cell(row=r, column=3, value=len(sd.cell_differences))
            ws.cell(
                row=r,
                column=4,
                value=str(sd.expected_shape) if sd.expected_shape else "-",
            )
            ws.cell(
                row=r,
                column=5,
                value=str(sd.actual_shape) if sd.actual_shape else "-",
            )
            ws.cell(row=r, column=6, value=notes)

        # Widen columns for readability
        widths = [32, 36, 14, 20, 20, 60]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _summarize_sheet(
        self, sd: SheetDiff
    ) -> tuple[str, str, PatternFill | None]:
        if sd.missing_in_actual:
            return (
                "Missing in actual / 実データ欠落",
                "Sheet exists in expected only / 期待値にのみ存在",
                FILL_MISSING,
            )
        if sd.missing_in_expected:
            return (
                "Missing in expected / 期待値欠落",
                "Sheet exists in actual only / 実データにのみ存在",
                FILL_MISSING,
            )
        if sd.read_error:
            return ("Read error / 読み込みエラー", sd.read_error, FILL_STRUCTURAL)

        notes: list[str] = []
        fill: PatternFill | None = None

        if not sd.shape_matches:
            notes.append("Shape mismatch / サイズ不一致")
            fill = FILL_STRUCTURAL
        if not sd.columns_match:
            notes.append("Column names differ / 列名不一致")
            fill = FILL_STRUCTURAL
        if sd.cell_differences:
            notes.append(
                f"{len(sd.cell_differences)} cell diff(s) / {len(sd.cell_differences)} 件のセル差分"
            )
            if fill is None:
                fill = FILL_DIFF

        if not notes:
            return ("✓ Match / 一致", "-", None)
        return ("✗ Differences / 差分あり", "; ".join(notes), fill)

    # --- Excel content sheets -----------------------------------------------

    def _write_excel_content_sheets(self, wb: Workbook) -> None:
        """For each sheet common to both files, render the actual data with diffs highlighted."""
        actual_path = Path(self.result.actual_path)
        try:
            actual_xl = pd.ExcelFile(actual_path)
        except Exception as e:
            logger.warning("Could not reopen actual Excel for highlighting: %s", e)
            return

        sheet_diff_by_name = {sd.sheet_name: sd for sd in self.result.sheet_diffs}

        for name in actual_xl.sheet_names:
            sd = sheet_diff_by_name.get(str(name))
            if sd is None or sd.missing_in_expected:
                # Sheet only in actual — render as-is, marked missing_in_expected.
                df = pd.read_excel(actual_xl, sheet_name=name)
                ws = wb.create_sheet(self._safe_sheet_name(str(name)))
                self._render_dataframe(ws, df, highlights={}, full_row_highlight=True)
                continue
            if sd.missing_in_actual:
                # Nothing to render from actual; summary covers it.
                continue

            df = pd.read_excel(actual_xl, sheet_name=name)
            highlights = self._build_highlight_map(sd)
            ws = wb.create_sheet(self._safe_sheet_name(str(name)))
            self._render_dataframe(ws, df, highlights=highlights)

    # --- CSV content sheet --------------------------------------------------

    def _write_csv_content_sheet(self, wb: Workbook) -> None:
        if not self.result.sheet_diffs:
            return
        sd = self.result.sheet_diffs[0]
        actual_path = Path(self.result.actual_path)
        try:
            df = _read_csv_best_effort(actual_path)
        except Exception as e:
            logger.warning("Could not reopen actual CSV for highlighting: %s", e)
            return

        ws = wb.create_sheet(self._safe_sheet_name(sd.sheet_name))
        highlights = self._build_highlight_map(sd)
        self._render_dataframe(ws, df, highlights=highlights)

    # --- Rendering helpers --------------------------------------------------

    def _build_highlight_map(
        self, sd: SheetDiff
    ) -> dict[tuple[int, str], str]:
        """Map (data_row_idx, column_name) → comment text with the expected value."""
        mapping: dict[tuple[int, str], str] = {}
        for cd in sd.cell_differences:
            mapping[(cd.row, cd.column)] = (
                f"Expected / 期待値: {_format_for_comment(cd.expected_value)}\n"
                f"Actual / 実際値: {_format_for_comment(cd.actual_value)}"
            )
        return mapping

    def _render_dataframe(
        self,
        ws: Worksheet,
        df: pd.DataFrame,
        highlights: dict[tuple[int, str], str],
        full_row_highlight: bool = False,
    ) -> None:
        # Header row
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = Alignment(vertical="center")

        # Data rows
        for row_idx in range(len(df)):
            excel_row = row_idx + 2  # 1-based + header
            for col_idx, col_name in enumerate(df.columns, start=1):
                raw = df.iloc[row_idx, col_idx - 1]
                value = None if pd.isna(raw) else _excel_safe(raw)
                cell = ws.cell(row=excel_row, column=col_idx, value=value)

                if full_row_highlight:
                    cell.fill = FILL_MISSING
                    continue

                comment_text = highlights.get((row_idx, str(col_name)))
                if comment_text is not None:
                    cell.fill = FILL_DIFF
                    cell.comment = Comment(comment_text, "excel-diff-extractor")

        # Reasonable column widths
        for col_idx, col_name in enumerate(df.columns, start=1):
            header_len = len(str(col_name))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max(header_len + 2, 12), 40
            )

        ws.freeze_panes = "A2"

    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        """Excel sheet names are capped at 31 chars and forbid [ ] : * ? / \\ .

        We replace forbidden characters with '_' rather than dropping them so
        names like "Q1/Q2 Summary" stay readable as "Q1_Q2 Summary" instead of
        collapsing to "Q1Q2 Summary".
        """
        cleaned = "".join("_" if c in r"[]:*?/\\" else c for c in name)
        return cleaned[:31] or "Sheet"


# --- Private helpers -------------------------------------------------------


def _excel_safe(value: Any) -> Any:
    """openpyxl can't write arbitrary numpy / pandas types — coerce them."""
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    try:
        import numpy as np  # local import keeps openpyxl-only users lean

        if isinstance(value, np.integer):
            return int(value)
        if isinstance(value, np.floating):
            return float(value)
        if isinstance(value, np.bool_):
            return bool(value)
    except ImportError:
        pass
    if isinstance(value, (int, float, str, bool)):
        return value
    return str(value)


def _format_for_comment(value: Any) -> str:
    """Human-readable representation of a cell value for a diff comment.

    Avoids exposing numpy type repr like 'np.int64(200)' — users don't care
    about the underlying dtype, they want to see '200'.
    """
    if pd.isna(value):
        return "(empty / 空)"
    clean = _excel_safe(value)
    if isinstance(clean, float):
        # Trim trailing zeros for readability while preserving precision.
        text = f"{clean:.10g}"
        return text
    if isinstance(clean, str):
        return f'"{clean}"'
    return str(clean)


def _read_csv_best_effort(path: Path) -> pd.DataFrame:
    for encoding in ("utf-8", "cp932", "shift_jis", "latin-1"):
        try:
            return pd.read_csv(path, encoding=encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Unable to decode CSV: {path}")
