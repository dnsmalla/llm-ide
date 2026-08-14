"""End-to-end regression test: synthetic 2-product plan, full pipeline."""

import json
from pathlib import Path

import pytest

gp = pytest.importorskip("gurobipy")
from openpyxl import Workbook

from result_explanation.analyzer import AnalysisOptions, run_analysis


@pytest.fixture()
def fixture_dir(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Capacity"
    ws.append(["Resource", "Hours available"])
    ws.append(["machine", 80])
    ws2 = wb.create_sheet("Demand")
    ws2.append(["Product", "Max demand", "Profit per unit", "Hours per unit"])
    ws2.append(["A", 50, 3, 1.0])
    ws2.append(["B", 60, 2, 0.5])
    wb.save(tmp_path / "input.xlsx")

    env = gp.Env(empty=True)
    env.setParam("OutputFlag", 0)
    env.start()
    m = gp.Model("plan", env=env)
    x_a = m.addVar(name="make_A", lb=0)
    x_b = m.addVar(name="make_B", lb=0)
    m.addConstr(1.0 * x_a + 0.5 * x_b <= 80, name="machine_hours")
    m.addConstr(x_a <= 50, name="demand_A")
    m.addConstr(x_b <= 60, name="demand_B")
    m.setObjective(3 * x_a + 2 * x_b, gp.GRB.MAXIMIZE)
    m.write(str(tmp_path / "model.lp"))
    m.optimize()

    wb = Workbook()
    ws = wb.active
    ws.append(["Variable", "Value"])
    ws.append(["make_A", x_a.X])
    ws.append(["make_B", x_b.X])
    ws.append(["Total profit", m.ObjVal])
    wb.save(tmp_path / "output.xlsx")
    return tmp_path


def _run(fixture_dir: Path, **kwargs) -> dict:
    options = AnalysisOptions(
        input_excels=[fixture_dir / "input.xlsx"],
        output_excels=[fixture_dir / kwargs.pop("output_name", "output.xlsx")],
        output_dir=fixture_dir / "out",
        **kwargs,
    )
    report = run_analysis(fixture_dir / "model.lp", options)
    facts = json.loads((fixture_dir / "out" / "model_facts.json").read_text())
    return {"report": report, "facts": facts}


def test_consistent_run_explained(fixture_dir: Path):
    result = _run(fixture_dir, resolve=True, focus=["make_A"])
    facts = result["facts"]
    assert facts["meta"]["coverage_pct"] == 100.0
    assert facts["binding_summary"]["n_binding"] == 3
    assert facts["binding_summary"]["n_violated"] == 0
    assert not facts["consistency"]["objective_mismatch"]
    assert facts["consistency"]["optimality_gap_vs_resolve"] == pytest.approx(0.0)
    # demand_A RHS traced to the input workbook
    make_a = next(d for d in facts["drivers"] if d["variable"] == "make_A")
    demand_a = next(c for c in make_a["binding_constraints"] if c["constraint"] == "demand_A")
    assert demand_a["rhs_source"]["found"]
    assert demand_a["rhs_source"]["cells"][0]["sheet"] == "Demand"
    assert demand_a["shadow_price"] == pytest.approx(3.0)
    # report contains the splice placeholder
    text = result["report"].read_text()
    assert "produced by the summarizer subagent" in text


def test_resolve_attaches_rhs_valid_range(fixture_dir: Path):
    """--resolve surfaces the RHS sensitivity range alongside the shadow price."""
    facts = _run(fixture_dir, resolve=True, focus=["make_A"])["facts"]
    make_a = next(d for d in facts["drivers"] if d["variable"] == "make_A")
    demand_a = next(c for c in make_a["binding_constraints"] if c["constraint"] == "demand_A")
    assert "shadow_price" in demand_a
    lo, hi = demand_a["rhs_valid_range"]
    assert lo <= demand_a["rhs"] <= hi


def test_reported_optimal_flagged_when_matches(fixture_dir: Path):
    """When the reported solution IS Gurobi's optimum, the duals describe it."""
    result = _run(fixture_dir, resolve=True)
    facts = result["facts"]
    assert facts["consistency"]["reported_is_optimal"] is True
    assert facts["meta"]["duals_describe_reported"] is True


def test_suboptimal_reported_marks_duals_as_optimum_only(fixture_dir: Path):
    """A feasible-but-suboptimal reported solution must NOT claim the optimum's
    duals describe it — that is the misleading case we guard against."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Variable", "Value"])
    ws.append(["make_A", 40.0])  # feasible (hours 60 <= 80) but obj 200 < 270
    ws.append(["make_B", 40.0])
    ws.append(["Total profit", 200.0])
    wb.save(fixture_dir / "output_subopt.xlsx")
    facts = _run(fixture_dir, output_name="output_subopt.xlsx", resolve=True)["facts"]
    assert facts["binding_summary"]["n_violated"] == 0
    assert facts["consistency"]["reported_is_optimal"] is False
    assert facts["meta"]["duals_describe_reported"] is False
    gap = facts["consistency"]["optimality_gap_vs_resolve"]
    assert gap == pytest.approx(200.0 - 270.0)


def test_violated_output_flagged(fixture_dir: Path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Variable", "Value"])
    ws.append(["make_A", 70.0])  # violates demand_A and machine_hours
    ws.append(["make_B", 60.0])
    wb.save(fixture_dir / "output_bad.xlsx")
    facts = _run(fixture_dir, output_name="output_bad.xlsx")["facts"]
    violated = {v["name"] for v in facts["violated_constraints"]}
    assert violated == {"machine_hours", "demand_A"}


def test_corrupt_input_excel_does_not_abort(fixture_dir: Path):
    """A bad input workbook is skipped with a warning; the run still completes."""
    bad = fixture_dir / "corrupt.xlsx"
    bad.write_text("not really a workbook")
    options = AnalysisOptions(
        input_excels=[bad, fixture_dir / "input.xlsx"],
        output_excels=[fixture_dir / "output.xlsx"],
        output_dir=fixture_dir / "out",
    )
    report = run_analysis(fixture_dir / "model.lp", options)
    assert report.exists()


def test_cli_invalid_model_exits_cleanly(tmp_path: Path):
    """Feeding a non-model file returns a non-zero code, not a traceback."""
    from result_explanation.cli import main

    junk = tmp_path / "junk.lp"
    junk.write_text("this is not a valid lp model @@@@\n")
    assert main([str(junk)]) != 0


def test_sol_file_preferred(fixture_dir: Path):
    (fixture_dir / "model.sol").write_text("# Objective value = 270\nmake_A 50\nmake_B 60\n")
    facts = _run(fixture_dir, sol_file=fixture_dir / "model.sol")["facts"]
    assert facts["meta"]["coverage_pct"] == 100.0
    assert facts["consistency"]["reported_objective"] == pytest.approx(270.0)
    assert facts["binding_summary"]["n_violated"] == 0


def _index_model(*names: str) -> gp.Model:
    env = gp.Env(empty=True)
    env.setParam("OutputFlag", 0)
    env.start()
    m = gp.Model("idx", env=env)
    for name in names:
        m.addVar(name=name, lb=0)
    m.update()
    return m


def test_index_ranges_zero_based():
    """コマが 0 始まりなら has_negative_index は False。"""
    from result_explanation import lp_extract

    m = _index_model("v_p_hp(0_0)", "v_p_hp(3_47)", "total_pump_power(0)")
    ranges = lp_extract.index_ranges(m)
    assert ranges["global_min_index"] == 0
    assert ranges["has_negative_index"] is False
    # 位置 0: ユニット 0..3、位置 1: コマ 0..47
    pos = {p["position"]: p for p in ranges["positions"]}
    assert pos[0]["min"] == 0 and pos[0]["max"] == 3
    assert pos[1]["min"] == 0 and pos[1]["max"] == 47


def test_index_ranges_detects_minus_one():
    """初期コマを -1 で持つモデルは global_min_index が -1。"""
    from result_explanation import lp_extract

    m = _index_model("level(-1_0)", "level(0_0)", "level(0_47)")
    ranges = lp_extract.index_ranges(m)
    assert ranges["global_min_index"] == -1
    assert ranges["has_negative_index"] is True


def test_index_ranges_negative_but_not_minus_one():
    """-2 始まりでも has_negative_index は True、起点判定は global_min_index で行う。"""
    from result_explanation import lp_extract

    m = _index_model("level(-2_0)", "level(0_0)", "level(0_47)")
    ranges = lp_extract.index_ranges(m)
    assert ranges["global_min_index"] == -2
    assert ranges["has_negative_index"] is True


def test_index_ranges_empty_when_unindexed():
    """添字を持たない変数のみなら空 dict。"""
    from result_explanation import lp_extract

    m = _index_model("make_A", "make_B")
    assert lp_extract.index_ranges(m) == {}


_SUMMARY = "## Verdict\n\nPumping is zeroed by cost.\n\n## Why\n\n…\n\n## What Would Change It\n\n…"


def test_rerun_preserves_spliced_summary(fixture_dir: Path):
    """エンジン再実行でスプライス済みサマリが消えず引き継がれる。"""
    from result_explanation import i18n
    from result_explanation.report_generator import PLACEHOLDER

    note = i18n.tr("en", "regen_note")

    # 1 回目の実行 → プレースホルダ入りレポート
    report = _run(fixture_dir, sol_file=_write_sol(fixture_dir))["report"]
    assert PLACEHOLDER in report.read_text()

    # オーケストレータによるスプライスを模擬
    report.write_text(report.read_text().replace(PLACEHOLDER, _SUMMARY))

    # 2 回目の実行（検証/再実行を模擬）→ サマリは保持され注記が付く
    report2 = _run(fixture_dir, sol_file=fixture_dir / "model.sol")["report"]
    text = report2.read_text()
    assert "Pumping is zeroed by cost." in text
    assert PLACEHOLDER not in text
    assert note in text

    # 3 回目でも注記は重複しない
    report3 = _run(fixture_dir, sol_file=fixture_dir / "model.sol")["report"]
    assert report3.read_text().count(note) == 1


def _min_facts() -> dict:
    return {
        "meta": {
            "n_vars": 1,
            "n_linear_constrs": 0,
            "is_mip": False,
            "obj_sense": "minimize",
            "solution_source": "x",
            "n_values": 1,
            "coverage_pct": 100.0,
        },
        "consistency": {"reported_objective": None, "objective_mismatch": False},
        "objective_breakdown": {"computed_objective": 0.0, "families": [], "n_families_total": 0},
        "binding_summary": {"n_binding": 0, "n_slack": 0, "n_violated": 0, "n_not_evaluable": 0},
        "violated_constraints": [],
        "drivers": [],
        "other_binding_constraints": [],
    }


def test_report_has_collapsible_and_markers():
    """技術詳細は折りたたみ <details> に入り、サマリはマーカーで囲まれる。"""
    from result_explanation import i18n
    from result_explanation.report_generator import SUMMARY_END, SUMMARY_START, render

    text = render(_min_facts(), Path("m.lp"), "en")
    assert "<details>" in text and "</details>" in text
    assert SUMMARY_START in text and SUMMARY_END in text
    # 折りたたみ summary に技術詳細ラベル、Run summary は details の内側
    assert i18n.tr("en", "tech_details_summary") in text
    assert text.index("<details>") < text.index(i18n.tr("en", "run_summary_header"))


def test_japanese_report():
    """language='ja' で見出しが日本語になる。"""
    from result_explanation.report_generator import render

    text = render(_min_facts(), Path("m.lp"), "ja")
    assert "結果説明レポート" in text
    assert "実行サマリ" in text
    assert "目的関数の内訳（変数ファミリ別）" in text


def test_default_output_dir_under_debug():
    """lp が debug/<step>/ 配下なら出力は debug/ 直下の result_explanation。"""
    from result_explanation.analyzer import _default_output_dir

    lp = Path("/run/mode_all_split/debug/step3_simple/step3_simple_model.lp")
    assert _default_output_dir(lp) == Path("/run/mode_all_split/debug/result_explanation")
    # debug が無ければ従来どおりモデル隣
    lp2 = Path("/tmp/foo/model.lp")
    assert _default_output_dir(lp2) == Path("/tmp/foo/result_explanation")


def test_extract_explanation_ignores_placeholder():
    """プレースホルダのままなら extract_explanation は None。"""
    from result_explanation.report_generator import PLACEHOLDER, extract_explanation, render

    text = render(_min_facts(), Path("m.lp"))
    assert PLACEHOLDER in text
    assert extract_explanation(text) is None


def _write_sol(fixture_dir: Path) -> Path:
    sol = fixture_dir / "model.sol"
    sol.write_text("# Objective value = 270\nmake_A 50\nmake_B 60\n")
    return sol
