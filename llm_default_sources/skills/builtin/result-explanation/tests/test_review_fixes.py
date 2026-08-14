"""Regression tests for the production-review findings:
ranged constraints, rounding-aware violation/at-bound tolerances, objective pick."""

import pytest

gp = pytest.importorskip("gurobipy")

from result_explanation import binding, solution_reader  # noqa: E402


def _model(build) -> gp.Model:
    env = gp.Env(empty=True)
    env.setParam("OutputFlag", 0)
    env.start()
    m = gp.Model(env=env)
    build(m)
    m.update()
    return m


# ── C1: ranged constraints are evaluated against [lo, hi], not dropped ──

def test_ranged_constraint_binding_not_unknown():
    def build(m):
        x = m.addVar(lb=0, ub=100, name="x")
        m.addRange(1.0 * x, 10.0, 50.0, name="band")  # 10 <= x <= 50
        m.setObjective(x, gp.GRB.MAXIMIZE)
    m = _model(build)
    # x at the upper edge of the band → binding, not "not evaluable".
    res = binding.analyze_constraints(m, {"x": 50.0})
    assert res["n_unknown"] == 0
    assert any(e["name"] == "band" for e in res["binding"])


def test_ranged_constraint_violation_detected():
    def build(m):
        x = m.addVar(lb=0, ub=100, name="x")
        m.addRange(1.0 * x, 10.0, 50.0, name="band")
        m.setObjective(x, gp.GRB.MAXIMIZE)
    m = _model(build)
    res = binding.analyze_constraints(m, {"x": 80.0})  # above hi=50
    assert any(e["name"] == "band" for e in res["violated"])


def test_user_var_named_rg_not_mistaken_for_range_aux():
    """REGRESSION: a user variable like 'Rgn' must NOT be treated as a range
    auxiliary; its value must count toward activity (so violations show)."""
    def build(m):
        rg = m.addVar(lb=0, ub=100, name="Rgn")  # legitimate user variable
        y = m.addVar(lb=0, ub=100, name="y")
        m.addConstr(rg + y <= 50, name="cn")      # plain <= constraint, NOT ranged
        m.setObjective(rg + y, gp.GRB.MAXIMIZE)
    m = _model(build)
    # Rgn=45, y=10 → activity 55 > 50 ⇒ must be VIOLATED, not silently "satisfied".
    res = binding.analyze_constraints(m, {"Rgn": 45.0, "y": 10.0})
    assert any(e["name"] == "cn" for e in res["violated"])
    assert res["n_slack"] == 0


def test_ranged_constraint_interior_is_slack():
    def build(m):
        x = m.addVar(lb=0, ub=100, name="x")
        m.addRange(1.0 * x, 10.0, 50.0, name="band")
        m.setObjective(x, gp.GRB.MAXIMIZE)
    m = _model(build)
    res = binding.analyze_constraints(m, {"x": 30.0})  # inside the band
    assert res["n_slack"] == 1
    assert not res["binding"] and not res["violated"]


# ── C2: a value rounded to 2 decimals must NOT be flagged violated ──

def test_rounded_equality_not_falsely_violated():
    def build(m):
        x = m.addVar(lb=0, ub=100, name="x")
        m.addConstr(3.0 * x == 100.0, name="eq")  # exact x = 33.3333…
        m.setObjective(x)
    m = _model(build)
    # Excel reports x rounded to 33.33 → activity 99.99, off by 0.01.
    res = binding.analyze_constraints(m, {"x": 33.33})
    assert not res["violated"], "rounding must not register as a violation"
    assert any(e["name"] == "eq" for e in res["binding"])


def test_real_violation_still_detected():
    def build(m):
        x = m.addVar(lb=0, ub=100, name="x")
        m.addConstr(x <= 50.0, name="cap")
        m.setObjective(x)
    m = _model(build)
    res = binding.analyze_constraints(m, {"x": 70.0})  # 20 over — real violation
    assert any(e["name"] == "cap" for e in res["violated"])


# ── I2: at-bound detection is rounding-aware ──

def test_at_bound_rounding_aware():
    def build(m):
        m.addVar(lb=0, ub=500.0, name="x")
    m = _model(build)
    st = binding.variable_status(m, {"x": 499.97})  # rounded, effectively at ub
    assert st["x"]["at"] == "upper_bound"


def test_interior_not_flagged_at_bound():
    def build(m):
        m.addVar(lb=0, ub=500.0, name="x")
    m = _model(build)
    st = binding.variable_status(m, {"x": 250.0})
    assert st["x"]["at"] == "interior"


# ── I3: reported-objective picks the value column, not a leading year/index ──

def test_find_reported_objective_skips_leading_year(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["総費用", 2024, 1500000])  # label, year, actual value
    p = tmp_path / "out.xlsx"
    wb.save(p)
    assert solution_reader.find_reported_objective([p]) == 1500000


# ── I4: fuzzy token collisions are contested, not silently mis-assigned ──

def test_fuzzy_token_collision_dropped(tmp_path):
    from openpyxl import Workbook
    # p_hp(0_1) and p_hp(1_0) have identical token sets {p,hp,0,1}.
    wb = Workbook()
    ws = wb.active
    ws.title = "vals"
    ws.append(["", "p hp 0 1"])     # header whose tokens are the shared set
    ws.append(["row", 42.0])
    p = tmp_path / "piv.xlsx"
    wb.save(p)
    got = solution_reader.extract_fuzzy([p], {"p_hp(0_1)", "p_hp(1_0)"})
    # both collide on the same cell → neither is assigned (contested)
    assert "p_hp(0_1)" not in got and "p_hp(1_0)" not in got


def test_fuzzy_unique_match_resolves(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "prod"
    ws.append(["", "plant7"])
    ws.append(["row", 88.0])
    p = tmp_path / "piv2.xlsx"
    wb.save(p)
    got = solution_reader.extract_fuzzy([p], {"prod(plant7)", "prod(plant9)"})
    assert got == {"prod(plant7)": 88.0}  # only plant7 matches the cell


# ── M1: SAObj objective-coefficient range threaded into driver facts ──

def test_obj_coeff_range_in_facts(tmp_path):
    from result_explanation import facts as facts_mod
    var_status = {"x": {"value": 5.0, "lb": 0.0, "ub": None, "at": "interior",
                        "obj_coeff": 3.0, "vtype": "C"}}
    f = facts_mod.build_facts(
        meta={}, consistency={},
        objective={"computed_objective": 15.0, "families": [], "n_families_total": 0},
        drivers=["x"], var_status=var_status,
        constr_analysis={"binding": [], "violated": [], "n_slack": 0, "n_unknown": 0},
        duals={}, reduced_costs={}, value_index={},
        obj_ranges={"x": (1.0, 7.0)},
    )
    assert f["drivers"][0]["obj_coeff_range"] == [1.0, 7.0]


# ── I1: analysis-confidence reflects coverage / non-evaluable / fuzzy ──

def test_analysis_confidence_levels():
    from result_explanation.analyzer import _analysis_confidence
    full = {"n_unknown": 0}
    partial = {"n_unknown": 8}
    # full coverage, exact source → high
    assert _analysis_confidence(100.0, full, 10, ".sol file (10 vars)")["level"] == "high"
    # fuzzy source → medium even at full coverage
    assert _analysis_confidence(100.0, full, 10, "output Excel fuzzy match")["level"] == "medium"
    # most constraints not evaluable → low, with a reason
    low = _analysis_confidence(40.0, partial, 10, "output Excel exact match")
    assert low["level"] == "low"
    assert low["reasons"]
