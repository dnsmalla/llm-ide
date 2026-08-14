#!/usr/bin/env python3
"""
LP File Comparator — standalone script for comparing two LP (Linear Programming) files.

Parses objectives and constraints, normalizes terms, handles SOS2/piecewise constraints,
and outputs a structured summary of all differences.

Usage:
    python lp_comparator.py <file1.lp> <file2.lp> [--label1 NAME] [--label2 NAME] [--csv PATH] [--json]
"""

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

try:
    import pandas as pd

    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# SOS Constraint Parsing & Comparison
# ---------------------------------------------------------------------------


class SOSConstraintParser:
    _sos2_re = re.compile(r"([-+]?\d+(?:\.\d+)?)\s+([^\s]*SOS2_y[^\s]*)")
    _rhs_re = re.compile(r"piecewise_hv[^=]*=\s*([-+]?\d+(?:\.\d+)?)")

    @classmethod
    def extract_values(cls, content: str) -> Dict[str, float]:
        values: Dict[str, float] = {}
        for line in content.split("\n"):
            if "SOS2_y" in line:
                m = cls._sos2_re.search(line)
                if m:
                    values[m.group(2)] = float(m.group(1))
            elif "piecewise_hv" in line and "=" in line:
                m = cls._rhs_re.search(line)
                if m:
                    values["RHS"] = float(m.group(1))
        return values


class SOSConstraintComparator:
    EPSILON = 1e-6

    @classmethod
    def _difference_entry(cls, val1: Any, val2: Any) -> Optional[Dict[str, Any]]:
        if isinstance(val1, (int, float)) and isinstance(val2, (int, float)):
            if abs(val1 - val2) < cls.EPSILON:
                return None
        elif val1 == val2:
            return None
        return {"debug_value": val1, "correct_value": val2}

    @classmethod
    def compare_constraints(cls, content1: str, content2: str) -> Tuple[bool, Dict[str, Any]]:
        values1 = SOSConstraintParser.extract_values(content1)
        values2 = SOSConstraintParser.extract_values(content2)

        differences = {}
        for var in set(values1.keys()) | set(values2.keys()):
            val1 = values1.get(var, "Not present")
            val2 = values2.get(var, "Not present")
            diff = cls._difference_entry(val1, val2)
            if diff is not None:
                differences[var] = diff

        return len(differences) == 0, differences


# ---------------------------------------------------------------------------
# Term Parsing & Normalization
# ---------------------------------------------------------------------------


class Term:
    __slots__ = ("sign", "coefficient", "variable")

    _extra_underscores_between_chars = re.compile(r"([^_])_+([^_])")
    _extra_underscores_before_paren = re.compile(r"_+\)")
    _spaces_before_paren = re.compile(r"\s*\(\s*")
    _spaces_after_paren = re.compile(r"\s*\)\s*")

    def __init__(self, sign: str, coefficient: str, variable: str):
        self.sign = sign if sign in ["+", "-"] else "+"
        self.coefficient = self._normalize_coefficient(coefficient)
        self.variable = self._normalize_variable(variable)

    @staticmethod
    def _normalize_coefficient(coef: str) -> str:
        if not coef or coef == ".":
            return "1"
        return coef

    @staticmethod
    def _normalize_variable(var: str) -> str:
        var = "".join(var.split())
        var = Term._extra_underscores_between_chars.sub(r"\1_\2", var)
        var = Term._extra_underscores_before_paren.sub(")", var)
        var = Term._spaces_before_paren.sub("(", var)
        var = Term._spaces_after_paren.sub(")", var)
        return var

    def __str__(self) -> str:
        return f"{self.sign}{self.coefficient}{self.variable}"


class TermParser:
    _term_pattern = re.compile(r"[+-]?\s*\d*\.?\d*\s*[a-zA-Z][a-zA-Z0-9_()]*")
    _term_parts_pattern = re.compile(r"([+-])?(\d*\.?\d*)?([a-zA-Z][a-zA-Z0-9_()]*)")

    @classmethod
    def parse_term(cls, term_str: str) -> Optional[Term]:
        term_str = term_str.strip()
        if not term_str:
            return None
        match = cls._term_parts_pattern.match(term_str)
        if not match:
            return None
        sign, coef, var = match.groups()
        return Term(sign or "+", coef, var)

    @classmethod
    def extract_terms(cls, expression: str) -> List[str]:
        return cls._term_pattern.findall(expression)


# ---------------------------------------------------------------------------
# LP Normalization
# ---------------------------------------------------------------------------


class LPNormalizer:
    OPERATORS = ["<=", ">=", "="]
    _operator_split_re = re.compile(r"(<=|>=|=)")

    @classmethod
    def normalize_term(cls, term: str) -> str:
        term = term.strip()
        if not term:
            return ""

        terms = TermParser.extract_terms(term)
        if not terms:
            return term

        normalized_terms = []
        for t in terms:
            if parsed_term := TermParser.parse_term(t):
                var_parts = parsed_term.variable.split("*")
                parsed_term.variable = "*".join(sorted(var_parts))
                normalized_terms.append(str(parsed_term))

        normalized_terms.sort(key=lambda x: (x[1:] if x.startswith("+") else x))

        rhs = ""
        for op in cls.OPERATORS:
            if op in term:
                parts = term.split(op)
                if len(parts) > 1:
                    rhs = op + parts[1].strip()
                break

        return "".join(normalized_terms) + rhs

    @classmethod
    def normalize_constraint(cls, constraint: str) -> str:
        if "SOS" in constraint or "piecewise" in constraint:
            values = SOSConstraintParser.extract_values(constraint)
            if values:
                return str(values)

        parts = cls._operator_split_re.split(constraint)
        if len(parts) < 3:
            return cls.normalize_term(constraint)

        left_side = parts[0]
        operator = parts[1]
        right_side = "".join(parts[2:])

        return f"{cls.normalize_term(left_side)}{operator}{cls.normalize_term(right_side)}"


# ---------------------------------------------------------------------------
# Comparators
# ---------------------------------------------------------------------------


class ObjectiveComparator:
    @staticmethod
    def _parse_terms(objective_str: str) -> Dict[str, str]:
        terms: Dict[str, str] = {}
        for line in objective_str.strip().split("\n"):
            line = line.strip()
            if not line:
                continue
            parts = line.split()
            if len(parts) < 2:
                continue
            try:
                coef = float(parts[0])
                var = " ".join(parts[1:])
                terms[var] = f"{coef} {var}"
            except (ValueError, IndexError):
                continue
        return terms

    def compare(self, obj1: str, obj2: str) -> Dict[str, Any]:
        if obj1 == obj2:
            return {"identical": True}
        obj1_terms = self._parse_terms(obj1)
        obj2_terms = self._parse_terms(obj2)
        all_vars = sorted(set(obj1_terms.keys()) | set(obj2_terms.keys()))
        differences = []
        for var in all_vars:
            t1 = obj1_terms.get(var, "")
            t2 = obj2_terms.get(var, "")
            if t1 != t2:
                differences.append({"variable": var, "file1_term": t1, "file2_term": t2})
        if not differences:
            return {"identical": True}
        return {
            "identical": False,
            "file1_objective": obj1,
            "file2_objective": obj2,
            "variable_differences": differences,
            "in_file1_only": [v for v in obj1_terms if v not in obj2_terms],
            "in_file2_only": [v for v in obj2_terms if v not in obj1_terms],
        }


class ConstraintComparator:
    def __init__(self, normalizer: LPNormalizer, is_sos_name):
        self._normalizer = normalizer
        self._is_sos_name = is_sos_name

    def compare(self, cons1: Dict[str, str], cons2: Dict[str, str]) -> Dict[str, Any]:
        names1 = set(cons1.keys())
        names2 = set(cons2.keys())

        if names1 == names2 and all(cons1[n] == cons2[n] for n in names1):
            return {
                "identical": True,
                "in_file1_only": {},
                "in_file2_only": {},
                "different_content": {},
                "common_count": len(names1),
            }

        in_file1_only: Dict[str, str] = {}
        in_file2_only: Dict[str, str] = {}
        different_content: Dict[str, Dict[str, Any]] = {}

        only1 = names1 - names2
        only2 = names2 - names1
        common = names1 & names2
        if only1:
            in_file1_only = {n: cons1[n] for n in sorted(only1)}
        if only2:
            in_file2_only = {n: cons2[n] for n in sorted(only2)}

        common_count = 0
        for name in sorted(common):
            c1 = cons1[name]
            c2 = cons2[name]
            if c1 == c2:
                common_count += 1
                continue
            if self._is_sos_name(name):
                identical, diffs = SOSConstraintComparator.compare_constraints(c1, c2)
                if identical:
                    common_count += 1
                else:
                    different_content[name] = {
                        "file1_content": c1,
                        "file2_content": c2,
                        "differences": diffs,
                    }
                continue
            if self._normalizer.normalize_constraint(c1) == self._normalizer.normalize_constraint(
                c2
            ):
                common_count += 1
            else:
                different_content[name] = {"file1_content": c1, "file2_content": c2}

        return {
            "identical": not in_file1_only and not in_file2_only and not different_content,
            "in_file1_only": in_file1_only,
            "in_file2_only": in_file2_only,
            "different_content": different_content,
            "common_count": common_count,
        }


# ---------------------------------------------------------------------------
# Constraint-family collapsing (for index-heavy Pyomo / Gurobi LP files)
# ---------------------------------------------------------------------------
#
# Pyomo emits one constraint per index, e.g. `pump_v_qw_cal_hp(0_0)`,
# `pump_v_qw_cal_hp(0_1)`, ... — so a name-by-name diff of two such models
# returns thousands of near-duplicate rows. Collapsing the *index* part of the
# name yields a "family" (`pump_v_qw_cal_hp`); counting families per file gives
# the structural diff a human actually wants: which constraint families were
# added / removed / changed in count between the two formulations.
#
# Only loop/time indices are collapsed; semantically meaningful numbers that
# are part of the name (e.g. `bigm1` vs `bigm2`, `com1`/`com2`/`com3`,
# `LEVEL4`, `gen1`) are preserved so distinct constraint types stay distinct.

_FAM_PAREN_IDX = re.compile(r"\([0-9_,]+\)")  # parenthesised index, e.g. (0_1)
_FAM_TIME_IDX = re.compile(r"_t\d+_?")  # time index, e.g. _t10_
_FAM_MID_IDX = re.compile(r"_\d+_")  # underscore-bounded index, _16_
_FAM_END_IDX = re.compile(r"_\d+$")  # trailing index, ..._8


def collapse_family(name: str) -> str:
    """Reduce an indexed constraint name to its index-free family name.

    Strips parenthesised indices and underscore-delimited loop/time indices,
    while keeping numbers fused to a word (bigm1, gen2) intact.
    """
    name = _FAM_PAREN_IDX.sub("", name)
    name = _FAM_TIME_IDX.sub("_t*_", name)
    prev = None
    while prev != name:  # repeat: adjacent indices like _1_2_ need two passes
        prev = name
        name = _FAM_MID_IDX.sub("_*_", name)
    name = _FAM_END_IDX.sub("_*", name)
    return name


# ---------------------------------------------------------------------------
# Main Comparator
# ---------------------------------------------------------------------------


class LPComparator:
    def __init__(self, file1_path: str, file2_path: str):
        self.file1_path = Path(file1_path)
        self.file2_path = Path(file2_path)
        self.file1_name = self.file1_path.name
        self.file2_name = self.file2_path.name
        self.lp1_content: Optional[str] = None
        self.lp2_content: Optional[str] = None
        self.lp1_objective: Optional[str] = None
        self.lp2_objective: Optional[str] = None
        self.lp1_constraints: Optional[Dict[str, str]] = None
        self.lp2_constraints: Optional[Dict[str, str]] = None
        self.normalizer = LPNormalizer()

    def read_files(self) -> bool:
        try:
            self.lp1_content = self.file1_path.read_text()
            self.lp2_content = self.file2_path.read_text()
            return True
        except Exception as e:
            print(f"Error reading files: {e}", file=sys.stderr)
            return False

    def _extract_objective(self, lp_content: str) -> str:
        objective_regex = re.compile(
            r"(?:Maximize|Minimize|max|min)(?:\s+obj:)?(.*?)(?:Subject To|s\.t\.)",
            re.DOTALL | re.IGNORECASE,
        )
        match = objective_regex.search(lp_content)
        if match:
            return match.group(1).strip()

        obj_regex = re.compile(r"obj:(.*?)(?:Subject To|s\.t\.)", re.DOTALL | re.IGNORECASE)
        match = obj_regex.search(lp_content)
        if match:
            return match.group(1).strip()

        lines = lp_content.split("\n")
        obj_lines = []
        capturing = False
        for line in lines:
            line = line.strip()
            if line.lower().startswith("obj:"):
                capturing = True
                continue
            elif capturing and line in ("Subject To", "s.t.", "Bounds", "End"):
                break
            elif capturing:
                obj_lines.append(line)
        return "\n".join(obj_lines)

    def _extract_constraints(self, lp_content: str) -> Dict[str, str]:
        constraints: Dict[str, str] = {}
        lines = lp_content.split("\n")
        current_constraint: List[str] = []
        current_name: Optional[str] = None
        in_constraints = False

        def save():
            nonlocal current_constraint, current_name
            if current_constraint and current_name:
                constraints[current_name] = "\n".join(current_constraint)
                current_constraint = []
                current_name = None

        for line in lines:
            line = line.strip()
            if not line:
                save()
                continue
            if line.startswith("Subject To") or line.startswith("s.t."):
                in_constraints = True
                continue
            elif line in ("Bounds", "Binary", "End", "Integer", "Generals"):
                save()
                in_constraints = False
                continue
            if not in_constraints:
                continue
            if ":" in line:
                save()
                current_name = line.split(":")[0].strip()
                current_constraint = [line.split(":", 1)[1].strip()]
            elif current_name:
                current_constraint.append(line)

        save()
        return constraints

    def parse_lp_files(self) -> bool:
        if not self.lp1_content or not self.lp2_content:
            if not self.read_files():
                return False
        if self.lp1_content is None or self.lp2_content is None:
            return False
        self.lp1_objective = self._extract_objective(self.lp1_content)
        self.lp1_constraints = self._extract_constraints(self.lp1_content)
        self.lp2_objective = self._extract_objective(self.lp2_content)
        self.lp2_constraints = self._extract_constraints(self.lp2_content)
        return True

    def compare(self) -> Dict[str, Any]:
        if not self.parse_lp_files():
            return {"error": "Failed to parse LP files"}
        obj_comp = ObjectiveComparator()
        con_comp = ConstraintComparator(self.normalizer, self._is_sos_name)
        return {
            "objective": obj_comp.compare(self.lp1_objective or "", self.lp2_objective or ""),
            "constraints": con_comp.compare(self.lp1_constraints or {}, self.lp2_constraints or {}),
        }

    @staticmethod
    def _is_sos_name(name: str) -> bool:
        return ("SOS" in name) or ("piecewise" in name)

    def get_summary_dataframe(self):
        if not HAS_PANDAS:
            return None
        differences = self.compare()
        if "error" in differences:
            print(f"Error: {differences['error']}", file=sys.stderr)
            return pd.DataFrame(
                columns=[
                    "section",
                    "type",
                    "variable",
                    f"{self.file1_name} (file1)",
                    f"{self.file2_name} (file2)",
                ]
            )
        builder = SummaryBuilder(self.file1_name, self.file2_name, self._is_sos_name)
        builder.add_objective_section(differences["objective"])
        builder.add_constraints_section(differences["constraints"])
        return builder.to_dataframe()

    def family_summary(self, filter_re: Optional[str] = None) -> Dict[str, Any]:
        """Group constraints into index-free families and count them per file.

        Returns a dict with a ``rows`` list (one entry per family) sorted with
        differences first. ``filter_re`` (case-insensitive regex) restricts the
        result to matching family names, e.g. ``"pump|_hp"``.
        """
        if not self.parse_lp_files():
            return {"error": "Failed to parse LP files"}

        cons1 = self.lp1_constraints or {}
        cons2 = self.lp2_constraints or {}
        fam1: "Counter[str]" = Counter(collapse_family(n) for n in cons1)
        fam2: "Counter[str]" = Counter(collapse_family(n) for n in cons2)

        pat = re.compile(filter_re, re.IGNORECASE) if filter_re else None
        families: Set[str] = set(fam1) | set(fam2)
        if pat:
            families = {f for f in families if pat.search(f)}

        rows: List[Dict[str, Any]] = []
        for fam in families:
            c1 = fam1.get(fam, 0)
            c2 = fam2.get(fam, 0)
            if c1 and not c2:
                status = "only_in_file1"
            elif c2 and not c1:
                status = "only_in_file2"
            elif c1 != c2:
                status = "count_differs"
            else:
                status = "same"
            rows.append({
                "family": fam,
                "count1": c1,
                "count2": c2,
                "diff": c2 - c1,
                "status": status,
            })

        rank = {"only_in_file2": 0, "only_in_file1": 1, "count_differs": 2, "same": 3}
        rows.sort(key=lambda r: (rank[str(r["status"])], str(r["family"])))
        return {"rows": rows, "filter": filter_re}

    def get_family_dataframe(self, filter_re: Optional[str] = None):
        if not HAS_PANDAS:
            return None
        summary = self.family_summary(filter_re)
        if "error" in summary:
            print(f"Error: {summary['error']}", file=sys.stderr)
            return None
        c1_col = f"{self.file1_name} (file1)"
        c2_col = f"{self.file2_name} (file2)"
        records = [
            {
                "constraint_family": r["family"],
                c1_col: r["count1"],
                c2_col: r["count2"],
                "diff(file2-file1)": r["diff"],
                "status": r["status"],
            }
            for r in summary["rows"]
        ]
        if not records:
            return pd.DataFrame(
                columns=["constraint_family", c1_col, c2_col, "diff(file2-file1)", "status"]
            )
        return pd.DataFrame(records)


# ---------------------------------------------------------------------------
# Summary Builder
# ---------------------------------------------------------------------------


class SummaryBuilder:
    def __init__(self, file1_name: str, file2_name: str, is_sos_name):
        self._rows: List[Dict[str, str]] = []
        self._file1_col = f"{file1_name} (file1)"
        self._file2_col = f"{file2_name} (file2)"
        self._is_sos_name = is_sos_name

    def _append(self, section, row_type, variable, debug_val, correct_val):
        self._rows.append({
            "section": section,
            "type": row_type,
            "variable": variable,
            self._file1_col: debug_val,
            self._file2_col: correct_val,
        })

    def add_objective_section(self, obj_diff: Dict[str, Any]):
        if obj_diff.get("identical", False):
            return
        for diff in obj_diff.get("variable_differences", []):
            self._append(
                "Objective",
                "Different Value",
                diff["variable"],
                diff["file1_term"] or "Not present",
                diff["file2_term"] or "Not present",
            )
        for var in obj_diff.get("in_file1_only", []):
            self._append(
                "Objective", "In File 1 Only", var, obj_diff["file1_objective"], "Not present"
            )
        for var in obj_diff.get("in_file2_only", []):
            self._append(
                "Objective", "In File 2 Only", var, "Not present", obj_diff["file2_objective"]
            )

    def add_constraints_section(self, const_diff: Dict[str, Any]):
        for name, content in const_diff.get("different_content", {}).items():
            if self._is_sos_name(name):
                self._add_sos_rows(name, content)

        non_sos_diff = [
            n for n in const_diff.get("different_content", {}) if not self._is_sos_name(n)
        ]
        only1 = {
            n: v for n, v in const_diff.get("in_file1_only", {}).items() if not self._is_sos_name(n)
        }
        only2 = {
            n: v for n, v in const_diff.get("in_file2_only", {}).items() if not self._is_sos_name(n)
        }
        all_non_sos = sorted(set(only1.keys()) | set(only2.keys()) | set(non_sos_diff))
        for name in all_non_sos:
            if name in only1:
                self._append("Constraints", "In File 1 Only", name, only1[name], "MISSING")
                continue
            if name in only2:
                self._append("Constraints", "In File 2 Only", name, "MISSING", only2[name])
                continue
            content = const_diff["different_content"][name]
            self._append(
                "Constraints",
                "Different Value",
                name,
                content["file1_content"],
                content["file2_content"],
            )

    def _add_sos_rows(self, name: str, content: Dict[str, Any]):
        if "differences" not in content:
            self._append(
                "Constraints",
                "Different Value (SOS)",
                name,
                content.get("file1_content", "Not present"),
                content.get("file2_content", "Not present"),
            )
            return
        diffs = content["differences"]
        if len(diffs) == 1:
            _, values = next(iter(diffs.items()))
            self._append(
                "Constraints",
                "Different Value (SOS)",
                name,
                str(values["debug_value"]),
                str(values["correct_value"]),
            )
            return
        debug_content = content.get("file1_content", "").strip()
        correct_content = content.get("file2_content", "").strip()
        if debug_content or correct_content:
            self._append(
                "Constraints",
                "Different Value (SOS)",
                name,
                debug_content or "Not present",
                correct_content or "Not present",
            )
            return
        debug_values = []
        correct_values = []
        for var, values in diffs.items():
            if values["debug_value"] != "Not present":
                debug_values.append(f"{var} = {values['debug_value']}")
            if values["correct_value"] != "Not present":
                correct_values.append(f"{var} = {values['correct_value']}")
        self._append(
            "Constraints",
            "Different Value (SOS)",
            name,
            "; ".join(debug_values) if debug_values else "Not present",
            "; ".join(correct_values) if correct_values else "Not present",
        )

    def to_dataframe(self):
        if not HAS_PANDAS:
            return None
        if not self._rows:
            return pd.DataFrame(
                columns=["section", "type", "variable", self._file1_col, self._file2_col]
            )
        return pd.DataFrame(self._rows)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def export_excel(df, output_path: str, file1_name: str, file2_name: str) -> str:
    if not HAS_OPENPYXL:
        csv_path = output_path.replace(".xlsx", ".csv")
        df.to_csv(csv_path, index=False)
        print("(Install openpyxl for colored Excel: pip install openpyxl)")
        return csv_path

    wb = Workbook()
    ws = wb.active
    ws.title = "LP Comparison"

    # --- Colors ---
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    DIFF_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # yellow
    FILE1_ONLY_FILL = PatternFill(
        start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"
    )  # red/pink
    FILE2_ONLY_FILL = PatternFill(
        start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"
    )  # green
    SOS_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # blue
    SECTION_FILL = PatternFill(
        start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
    )  # light grey
    THIN_BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    MISSING_FONT = Font(color="CC0000", italic=True, size=10)
    NORMAL_FONT = Font(size=10)
    BOLD_FONT = Font(bold=True, size=10)
    WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

    # --- Title row ---
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"LP Comparison: {file1_name} vs {file2_name}"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # --- Header row ---
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header.upper())
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 22

    # --- Data rows ---
    prev_section = None
    for row_idx, (_, row) in enumerate(df.iterrows(), 4):
        row_type = str(row.get("type", ""))
        section = str(row.get("section", ""))

        # Pick row fill based on difference type
        if "In File 1 Only" in row_type:
            fill = FILE1_ONLY_FILL
        elif "In File 2 Only" in row_type:
            fill = FILE2_ONLY_FILL
        elif "SOS" in row_type:
            fill = SOS_FILL
        elif "Different" in row_type:
            fill = DIFF_FILL
        else:
            fill = PatternFill()

        # Section separator
        if section != prev_section and prev_section is not None:
            for col_idx in range(1, len(headers) + 1):
                sep_cell = ws.cell(row=row_idx, column=col_idx)
                sep_cell.fill = SECTION_FILL
                sep_cell.border = THIN_BORDER
            # shift data one row down
            row_idx_actual = row_idx + 1
        else:
            row_idx_actual = row_idx

        for col_idx, header in enumerate(headers, 1):
            val = str(row[header]) if row[header] is not None else ""
            cell = ws.cell(row=row_idx_actual, column=col_idx, value=val)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN

            if val in ("MISSING", "Not present"):
                cell.font = MISSING_FONT
            elif header == "variable":
                cell.font = BOLD_FONT
            else:
                cell.font = NORMAL_FONT

        prev_section = section

    # --- Column widths ---
    col_widths = {"section": 14, "type": 22, "variable": 30}
    for col_idx, header in enumerate(headers, 1):
        width = col_widths.get(header, 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Legend sheet ---
    ws_legend = wb.create_sheet("Legend")
    legend_items = [
        ("Color", "Meaning"),
        ("Yellow", "Different Value — same constraint, different content"),
        ("Pink/Red", "In File 1 Only — missing from File 2"),
        ("Green", "In File 2 Only — missing from File 1"),
        ("Blue", "SOS/Piecewise — special constraint difference"),
    ]
    legend_fills = [
        HEADER_FILL,
        DIFF_FILL,
        FILE1_ONLY_FILL,
        FILE2_ONLY_FILL,
        SOS_FILL,
    ]
    legend_fonts = [
        HEADER_FONT,
        NORMAL_FONT,
        NORMAL_FONT,
        NORMAL_FONT,
        NORMAL_FONT,
    ]
    for i, (color_label, meaning) in enumerate(legend_items, 1):
        c1 = ws_legend.cell(row=i, column=1, value=color_label)
        c2 = ws_legend.cell(row=i, column=2, value=meaning)
        c1.fill = legend_fills[i - 1]
        c1.font = legend_fonts[i - 1]
        c2.font = legend_fonts[i - 1]
        if i == 1:
            c2.fill = HEADER_FILL
    ws_legend.column_dimensions["A"].width = 12
    ws_legend.column_dimensions["B"].width = 55

    wb.save(output_path)
    return output_path


def format_differences_text(differences: Dict[str, Any], file1_name: str, file2_name: str) -> str:
    lines = []

    obj = differences["objective"]
    if obj.get("identical"):
        lines.append("OBJECTIVE: Identical")
    else:
        lines.append("OBJECTIVE: Differences found")
        for d in obj.get("variable_differences", []):
            lines.append(
                f"  {d['variable']}: {d['file1_term'] or 'MISSING'} vs {d['file2_term'] or 'MISSING'}"
            )
        for v in obj.get("in_file1_only", []):
            lines.append(f"  {v}: only in {file1_name}")
        for v in obj.get("in_file2_only", []):
            lines.append(f"  {v}: only in {file2_name}")

    con = differences["constraints"]
    if con.get("identical"):
        lines.append(f"\nCONSTRAINTS: Identical ({con['common_count']} constraints)")
    else:
        lines.append(f"\nCONSTRAINTS: Differences found (common: {con['common_count']})")
        if con["in_file1_only"]:
            lines.append(
                f"  In {file1_name} only: {', '.join(sorted(con['in_file1_only'].keys()))}"
            )
        if con["in_file2_only"]:
            lines.append(
                f"  In {file2_name} only: {', '.join(sorted(con['in_file2_only'].keys()))}"
            )
        if con["different_content"]:
            lines.append(
                f"  Different content: {', '.join(sorted(con['different_content'].keys()))}"
            )
            for name, content in sorted(con["different_content"].items()):
                lines.append(f"\n  --- {name} ---")
                lines.append(f"  {file1_name}: {content['file1_content'][:200]}")
                lines.append(f"  {file2_name}: {content['file2_content'][:200]}")

    return "\n".join(lines)


def run_family_mode(comparator: "LPComparator", file1: Path, args) -> None:
    """Family-level comparison path (``--group-families``)."""
    summary = comparator.family_summary(args.filter)
    if "error" in summary:
        print(f"Error: {summary['error']}", file=sys.stderr)
        sys.exit(1)

    if args.json:
        print(json.dumps(summary, indent=2, default=str))
        return

    rows = summary["rows"]
    diff_rows = [r for r in rows if r["status"] != "same"]
    print("=" * 60)
    print("LP CONSTRAINT-FAMILY COMPARISON")
    print("=" * 60)
    print(f"File 1: {comparator.file1_name}")
    print(f"File 2: {comparator.file2_name}")
    if args.filter:
        print(f"Filter: {args.filter}")
    print(f"Families total: {len(rows)}  |  differing: {len(diff_rows)}")
    print()

    if not rows:
        print("No constraint families matched.")
        return

    if HAS_PANDAS:
        df = comparator.get_family_dataframe(args.filter)
        print(df.to_string(index=False))
        csv_path = args.csv or str(file1.parent / "lp_family_comparison.csv")
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"\nCSV saved to: {csv_path}")
    else:
        for r in rows:
            print(
                f"  {r['family']}: {r['count1']} vs {r['count2']} "
                f"({r['status']}, diff {r['diff']:+d})"
            )
        print("\n(Install pandas for a CSV export: pip install pandas)")


def main():
    parser = argparse.ArgumentParser(
        description="Compare two LP (Linear Programming) files and report differences.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Examples:\n"
        "  python lp_comparator.py model_debug.lp model_correct.lp\n"
        "  python lp_comparator.py a.lp b.lp --label1 debug --label2 correct\n"
        "  python lp_comparator.py a.lp b.lp --xlsx /path/to/output.xlsx\n"
        "  python lp_comparator.py a.lp b.lp --csv /path/to/output.csv\n"
        "  python lp_comparator.py a.lp b.lp --json",
    )
    parser.add_argument("file1", help="Path to the first LP file")
    parser.add_argument("file2", help="Path to the second LP file")
    parser.add_argument(
        "--label1", default=None, help="Label for file 1 in output (default: filename)"
    )
    parser.add_argument(
        "--label2", default=None, help="Label for file 2 in output (default: filename)"
    )
    parser.add_argument(
        "--xlsx", default=None, help="Path to save Excel output (default: auto-generated)"
    )
    parser.add_argument("--csv", default=None, help="Save as CSV instead of Excel")
    parser.add_argument("--json", action="store_true", help="Output raw JSON instead of table")
    parser.add_argument(
        "--group-families",
        action="store_true",
        help="Collapse Pyomo index suffixes into constraint families and compare "
        "per-family counts (use for large index-heavy LP files).",
    )
    parser.add_argument(
        "--filter",
        default=None,
        metavar="REGEX",
        help="With --group-families: keep only families matching this "
        "case-insensitive regex, e.g. 'pump|_hp'.",
    )
    args = parser.parse_args()

    file1 = Path(args.file1)
    file2 = Path(args.file2)

    if not file1.exists():
        print(f"Error: File not found: {file1}", file=sys.stderr)
        sys.exit(1)
    if not file2.exists():
        print(f"Error: File not found: {file2}", file=sys.stderr)
        sys.exit(1)

    comparator = LPComparator(str(file1), str(file2))
    if args.label1:
        comparator.file1_name = args.label1
    if args.label2:
        comparator.file2_name = args.label2

    if args.group_families:
        run_family_mode(comparator, file1, args)
        return

    differences = comparator.compare()

    if "error" in differences:
        print(f"Error: {differences['error']}", file=sys.stderr)
        sys.exit(1)

    if args.json:
        print(json.dumps(differences, indent=2, default=str))
        sys.exit(0)

    obj_identical = differences["objective"].get("identical", False)
    con_identical = differences["constraints"].get("identical", False)

    print("=" * 60)
    print("LP FILE COMPARISON REPORT")
    print("=" * 60)
    print(f"File 1: {file1.name}")
    print(f"File 2: {file2.name}")
    print()

    if obj_identical and con_identical:
        print("Result: FILES ARE IDENTICAL")
        print(f"  Constraints compared: {differences['constraints']['common_count']}")
        sys.exit(0)

    print(format_differences_text(differences, comparator.file1_name, comparator.file2_name))

    if HAS_PANDAS:
        df = comparator.get_summary_dataframe()
        if df is not None and not df.empty:
            print("\n" + "=" * 60)
            print("SUMMARY TABLE")
            print("=" * 60)
            print(df.to_string(index=False))

            if args.csv:
                df.to_csv(args.csv, index=False)
                print(f"\nCSV saved to: {args.csv}")
            else:
                xlsx_path = args.xlsx or str(file1.parent / "lp_comparison_result.xlsx")
                saved = export_excel(df, xlsx_path, comparator.file1_name, comparator.file2_name)
                print(f"\nExcel saved to: {saved}")
    else:
        print("\n(Install pandas for a formatted table: pip install pandas)")


if __name__ == "__main__":
    main()
