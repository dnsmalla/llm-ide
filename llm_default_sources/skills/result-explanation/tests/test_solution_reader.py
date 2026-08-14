"""Unit tests for solution extraction from output workbooks."""

from pathlib import Path

from openpyxl import Workbook

from result_explanation import solution_reader as sr


def test_tokenize_preserves_cjk():
    """CJK runs survive tokenization instead of collapsing to an empty set."""
    assert sr.tokenize("水力結果") == frozenset({"水力結果"})
    assert sr.tokenize("発電量 (MW)") == frozenset({"発電量", "mw"})
    # ASCII variable names tokenize exactly as before.
    assert sr.tokenize("v_p_hp(0_47)") == frozenset({"v", "p", "hp", "0", "47"})


def test_extract_exact_picks_value_column_by_header(tmp_path: Path):
    """With index columns before the value, the header keyword wins (not the
    first numeric cell, which would be an index)."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Variable", "unit", "step", "Value"])
    ws.append(["v_p_hp(0_47)", 0, 47, 123.5])  # first numeric is 0, value is 123.5
    p = tmp_path / "out.xlsx"
    wb.save(p)
    found, conflicts = sr.extract_exact([p], {"v_p_hp(0_47)"})
    assert found == {"v_p_hp(0_47)": 123.5}
    assert conflicts == []


def test_extract_exact_falls_back_to_last_numeric(tmp_path: Path):
    """No value-header: the LAST numeric cell is taken (indices sit to the left)."""
    wb = Workbook()
    ws = wb.active
    ws.append(["make_A", 0, 47, 88.0])
    p = tmp_path / "out.xlsx"
    wb.save(p)
    found, _ = sr.extract_exact([p], {"make_A"})
    assert found == {"make_A": 88.0}


def test_extract_exact_japanese_value_header(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.append(["変数", "コマ", "値"])
    ws.append(["make_A", 12, 50.0])
    p = tmp_path / "out.xlsx"
    wb.save(p)
    found, _ = sr.extract_exact([p], {"make_A"})
    assert found == {"make_A": 50.0}


def test_value_header_keywords_are_word_matched(tmp_path: Path):
    """A header like 'Max demand' must NOT be treated as a value column just
    because it contains the letter 'x'."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Variable", "Max demand", "Realized"])  # neither is a value col by word
    ws.append(["make_A", 50, 88.0])  # value should be the last numeric, 88.0
    p = tmp_path / "out.xlsx"
    wb.save(p)
    found, _ = sr.extract_exact([p], {"make_A"})
    assert found == {"make_A": 88.0}


def test_extract_exact_flags_conflict(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.append(["make_A", 50.0])
    ws.append(["make_A", 60.0])
    p = tmp_path / "out.xlsx"
    wb.save(p)
    found, conflicts = sr.extract_exact([p], {"make_A"})
    assert conflicts == ["make_A"]


def test_find_reported_objective_japanese(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.append(["目的関数値", 270.0])
    p = tmp_path / "out.xlsx"
    wb.save(p)
    assert sr.find_reported_objective([p]) == 270.0


def test_parse_sol_file_cp932(tmp_path: Path):
    """A .sol with a CP932-encoded comment line still parses values."""
    p = tmp_path / "m.sol"
    p.write_bytes("# 目的関数値\n# Objective value = 270\nmake_A 50\nmake_B 60\n".encode("cp932"))
    values, obj = sr.parse_sol_file(p)
    assert values == {"make_A": 50.0, "make_B": 60.0}
    assert obj == 270.0


def test_parse_sol_file_missing(tmp_path: Path):
    values, obj = sr.parse_sol_file(tmp_path / "nope.sol")
    assert values == {} and obj is None
